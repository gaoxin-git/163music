import sys
import datetime
import time
from selenium import webdriver

import urllib
import re
import requests

import openpyxl #excel文件读写xlsx

import multiprocessing

#使用前Internet/本地Internet/受信任的站定/受限制的站点中的启用保护模式全部去掉勾，或者全部勾上，否则报错
# options = webdriver.IeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-automation'])
# driverPath = r"E:\webdriverForSelenium\IEDriverServer.exe"

options = webdriver.ChromeOptions()

#模拟移动端访问
# 网易MV页面电脑端和移动端的内容不一样，移动端抓取元素更方便
# mobile_emulation = {"deviceName":"Google Nexus 5"}
# options.add_experimental_option("mobileEmulation", mobile_emulation)

# prefs = {
#     "download.prompt_for_download": False,
#     'download.default_directory': 'C:/Users/Administrator/Desktop/1/',#下载目录
#     "plugins.always_open_pdf_externally": True,
#     'profile.default_content_settings.popups': 0,#设置为0，禁止弹出窗口
#     # 'profile.default_content_setting_values.images': 2,#禁止图片加载
# }
# options.add_experimental_option('prefs', prefs)



driverPath = r"E:\webdriverForSelenium\chromedriver.exe"


HOME_PAGE = 'https://music.163.com/#/artist/mv?id=44266&limit=1000&offset=0'  #MV列表页面



keywords = []




# 添加系统环境变量
# os.environ['webdriver.ie.driver']= driverPath
# os.environ['webdriver.Chrome.driver']= driverPath

# browser = webdriver.Ie(executable_path=driverPath,options=options)

def read_articles(browser,page,file):
    """阅读文章"""

    articles = refresh_page(browser, page)
    length = len(articles)

    for index in range(length):
                # articles = browser.find_elements_by_xpath('//a[@class="fz14"]')
        try:
            print("共%s篇文章" % len(articles))
            article = articles[index]
            print('获取第%s篇文章：' % (index + 1))
            print(article.text)
            articls_info = []
            articls_info.append(article.text)

            article.click()
            #
            all_handles = browser.window_handles
            browser.switch_to_window(all_handles[-1])
            browser.get(browser.current_url)
        except:
            break

        try:
            # 发表时间
            try:
                publish_date = browser.find_element_by_xpath("//div[@class='head-tag']").text
                articls_info.append(publish_date[-19:-9])
            except:
                articls_info.append("")
            # author
            try:
                author = browser.find_element_by_xpath("//div[@class='author']").text
                articls_info.append(author)
            except:
                try:
                    author = browser.find_element_by_xpath("//div[@class='authorE']").text
                    articls_info.append(author)
                except:
                    continue
            # orgn
            try:
                orgn = browser.find_element_by_xpath("//div[@class='orgn']").text
                articls_info.append(orgn)
            except:
                orgn = browser.find_element_by_xpath("//div[@class='orgnE']").text
                articls_info.append(orgn)
            # ChDivSummary 摘要

            ChDivSummary = browser.find_element_by_xpath("//span[@id='ChDivSummary']")
            articls_info.append(ChDivSummary.text)
            # catalog_KEYWORD
            try:
                baseinfo = browser.find_element_by_xpath("//div[@class='wxBaseinfo']")
                baseinfo_subs = baseinfo.find_elements_by_xpath(".//p")
                #关键词
                catalog_KEYWORDs = baseinfo_subs[2].find_elements_by_xpath(".//a")

                articls_info.append("".join([k.text for k in catalog_KEYWORDs]))
            except:
                baseinfo = browser.find_element_by_xpath("//div[@class='wxInfo wxInfoEn']")
                baseinfo_subs = baseinfo.find_elements_by_xpath(".//p")
                #关键词
                catalog_KEYWORDs = baseinfo_subs[2].find_elements_by_xpath(".//a")

                articls_info.append("".join([k.text for k in catalog_KEYWORDs]))
            # sourinfo
            try:
                sourinfo = browser.find_element_by_xpath("//div[@class='sourinfo']")
                sourinfo_subs = sourinfo.find_elements_by_xpath(".//p")
                sourtitle = sourinfo_subs[0].text
                articls_info.append(sourtitle)
                sourtype = sourinfo_subs[-1].text
                articls_info.append(sourtype)
            except:
                articls_info.append("")
                articls_info.append("")

            write_excel_xlsx(file,'articles_summary',articls_info)

            print(articls_info)
        except:
            pass
        finally:
            browser.close()
            browser.switch_to_window(all_handles[0])
            time.sleep(1)
            articles = refresh_page(browser, page)
            while len(articles) == 0: #弹出验证码
                #_, articles = refresh_home(browser)
                #articles = refresh_page(browser, page)
                # 输入验证码
                # code = input("请输入验证码:")
                code = "quit"
                if code == "quit":  #有时候验证码正确也不能跳过验证页面，直接放弃
                    return
                code_edit = browser.find_element_by_xpath("//input[@id='CheckCode']")
                code_edit.click()
                code_edit.clear()
                code_edit.send_keys(code)
                code_edit = browser.find_element_by_xpath("//input[@type='button']")
                code_edit.click()



def read_articles_in_1st_page(browser,file):
    """阅读文章"""
    _,articles = refresh_home(browser,keywords)
    length = len(articles)

    for index in range(length):
                # articles = browser.find_elements_by_xpath('//a[@class="fz14"]')
        print("共%s篇文章" % len(articles))
        article = articles[index]
        print('获取第%s篇文章：' % (index + 1))
        print(article.text)
        articls_info = []
        articls_info.append(article.text)

        article.click()
        #
        all_handles = browser.window_handles
        browser.switch_to_window(all_handles[-1])
        browser.get(browser.current_url)

        try:
            # 发表时间
            try:
                publish_date = browser.find_element_by_xpath("//div[@class='head-tag']").text
                articls_info.append(publish_date[-19:-9])
            except:
                articls_info.append("")
            # author
            try:
                author = browser.find_element_by_xpath("//div[@class='author']").text
                articls_info.append(author)
            except:
                try:
                    author = browser.find_element_by_xpath("//div[@class='authorE']").text
                    articls_info.append(author)
                except:
                    continue
            # orgn
            try:
                orgn = browser.find_element_by_xpath("//div[@class='orgn']").text
                articls_info.append(orgn)
            except:
                orgn = browser.find_element_by_xpath("//div[@class='orgnE']").text
                articls_info.append(orgn)
            # ChDivSummary 摘要

            ChDivSummary = browser.find_element_by_xpath("//span[@id='ChDivSummary']")
            articls_info.append(ChDivSummary.text)
            # catalog_KEYWORD
            try:
                baseinfo = browser.find_element_by_xpath("//div[@class='wxBaseinfo']")
                baseinfo_subs = baseinfo.find_elements_by_xpath(".//p")
                #关键词
                catalog_KEYWORDs = baseinfo_subs[2].find_elements_by_xpath(".//a")

                articls_info.append("".join([k.text for k in catalog_KEYWORDs]))
            except:
                baseinfo = browser.find_element_by_xpath("//div[@class='wxInfo wxInfoEn']")
                baseinfo_subs = baseinfo.find_elements_by_xpath(".//p")
                #关键词
                catalog_KEYWORDs = baseinfo_subs[2].find_elements_by_xpath(".//a")

                articls_info.append("".join([k.text for k in catalog_KEYWORDs]))
            # sourinfo
            try:
                sourinfo = browser.find_element_by_xpath("//div[@class='sourinfo']")
                sourinfo_subs = sourinfo.find_elements_by_xpath(".//p")
                sourtitle = sourinfo_subs[0].text
                articls_info.append(sourtitle)
                sourtype = sourinfo_subs[-1].text
                articls_info.append(sourtype)
            except:
                articls_info.append("")
                articls_info.append("")

            write_excel_xlsx(file,'articles_summary',articls_info)

            print(articls_info)
        except:
            pass
        finally:
            browser.close()
            browser.switch_to_window(all_handles[0])
            time.sleep(3)
            _,articles = refresh_home(browser,keywords)
            if not articles:
                _,articles = refresh_home(browser,keywords)


def create_excel_xlsx(filename):
    # 创建工作簿
    workbook = openpyxl.Workbook()
    # 创建sheet
    data_sheet = workbook.create_sheet('articles_summary')
    row0 = [u'题目', u'发表时间', u'作者', u'机构',u'摘要',u'关键词',u'期刊题目','期刊类型']

    # 生成第一行和第二行
    for i in range(len(row0)):
        data_sheet.cell(1, i+1, row0[i])
    # 保存文件
    workbook.save(filename)

def write_excel_xlsx(filename, sheet_name, value):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    sheet.title = sheet_name
    rows_existed = sheet.max_row
    for j in range(0, len(value)):
        sheet.cell(row=1+rows_existed, column=j+1, value=str(value[j]))
    workbook.save(filename)


def refresh_home(browser,keywords):
    browser.get(HOME_PAGE)
    #输入搜索关键词
    try:
        search_edit = browser.find_elements_by_xpath('//input[@class="rekeyword"]')  #
        search_edit[0].click()
        search_edit[0].clear()
        search_edit[0].send_keys(keywords[0])
        # 点击搜索
        search_btn = browser.find_elements_by_xpath('//input[@class="researchbtn"]')  #
        search_btn[0].click()

        time.sleep(2)

        browser.switch_to.frame("iframeResult")
        time.sleep(1)
        # 获取文章列表
        articles = browser.find_elements_by_xpath('//a[@class="fz14"]')  #

        # TitleLeftCell
        pages_btn = browser.find_element_by_xpath('//table[@class="pageBar_bottom"]')  # 翻页
        pages_btn = pages_btn.find_elements_by_xpath('.//*[@href]')  # 翻页按钮  注意查找元素的子元素应使用  .// 而不是 //,否则就是在页面查找

        cnt_pages = len(pages_btn)
        print(cnt_pages)

        pages = []
        for pg in pages_btn:
            pages.append(pg.get_attribute('href'))
            print(pg.get_attribute('href'))

        return pages,articles
    except:
        return [],[]

def refresh_page(browser,page):
    browser.get(page)

    # time.sleep(2)
    # browser.switch_to.frame("iframeResult")
    # time.sleep(1)
    # 获取文章列表
    articles = browser.find_elements_by_xpath('//a[@class="fz14"]')  #
    return articles



def startSearch(keywords,textBrowser):
    # browser.maximize_window()
    global browser_main
    browser_main = webdriver.Chrome(executable_path=driverPath, options=options)

    pages,articls_in_1st_page = refresh_home(browser_main,keywords)

    length = len(pages)

    today = datetime.datetime.now()
    today = today.strftime("%Y%m%d-%H%M%S")
    filename = today + "".join(keywords) + ".xlsx"
    print(filename)

    textBrowser.append(filename)

    #打开excel文件
    create_excel_xlsx(filename)

    read_articles_in_1st_page(browser_main,filename)

    for page in pages[0:-1]:
        browser_main.close()
        browser_main = webdriver.Chrome(executable_path=driverPath, options=options)
        refresh_home(browser_main,keywords)
        read_articles(browser_main,page,filename)

def getMvUrl(url):
    title = url.text
    print(title)
    url = link.find_element_by_xpath('.//*[@href]')  #title
    url = url.get_attribute('href')  #url
    print(url)
    return (title,url)
def downloadMV(url,browser):  #url  :（text,hurl)
    # 新开一个窗口，通过执行js来新开一个窗口
    js = 'window.open("%s");' % url[1]
    browser.execute_script(js)

    # browser.get(url)

    browser.switch_to.window(browser.window_handles[-1])  # 切换新窗口标签
    browser.switch_to.frame("contentFrame")
    time.sleep(2)


    mvElement = browser.find_element_by_xpath('//div[@class="mv"]')#
    hurl = mvElement.get_attribute('data-flashvars')
    print(hurl)
    hurl = urllib.parse.unquote(hurl)  # url解码
    print(hurl)

    try:
        hurl = hurl.split('&')[0]
        hurl = hurl.split('hurl=')[-1]
        print('下载地址：%s' % hurl)

        download_music(url[0],hurl)
    except:

        pass
    finally:
        browser.close()
        browser.switch_to.window(browser.window_handles[0]) # 切换主标签

def download_music(title,url):
    print('{}正在下载'.format((url)))
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:63.0) Gecko/20100101 Firefox/63.0'}
    response = requests.get(url,headers=headers)
    with open('{}.mp4'.format(title), 'wb') as f:
        f.write(response.content)
        print('下载完成')

if __name__ == '__main__':
    browser = webdriver.Chrome(executable_path=driverPath, options=options)
    browser.get(HOME_PAGE)

    browser.switch_to.frame("contentFrame")
    time.sleep(1)
    # 获取文章列表
    mvlist = browser.find_element_by_xpath('//ul[@id="m-mv-module"]')  #
    mvlinks = mvlist.find_elements_by_xpath('.//p[@class="dec"]')

    print('共获取%d个MV资源' % len(mvlinks))

    mvUrls = []

    for i,link in enumerate(mvlinks):
        print('第%d个MV资源：' % i)
        mvUrls.append(getMvUrl(link))

    for i,url in enumerate(mvUrls):
        print('正在下载第%d个MV资源：' % i)
        downloadMV(url,browser)